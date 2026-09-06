"""Course/content master workbook that collects run outputs into curriculum-level records."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .naming import workbook_filename

ROOT = Path(__file__).resolve().parents[1]
MASTER_ROOT = ROOT / "output" / "curriculum_masters"
MASTER_CSV_DIR = MASTER_ROOT / "tables"
MASTER_XLSX = MASTER_ROOT / "curriculum_master.xlsx"

RUNS_HEADERS = [
    "job_id", "project_name", "course_name", "content_area", "subject_area", "specialty_area",
    "concept", "build_profile", "deck_style", "duplicate_policy", "duplicate_mode", "matched_job_id", "status",
    "batch_dir", "master_workbook",
]
OBJECTIVE_HEADERS = [
    "job_id", "project_name", "course_name", "content_area", "objective_id", "concept", "exemplar",
    "nclex_category", "ncjmm_primary", "priority_framework", "description",
]
CARD_HEADERS = [
    "job_id", "project_name", "course_name", "content_area", "card_id", "card_title", "concept",
    "subject_area", "specialty_area", "status", "evidence_status",
]
ARTIFACT_HEADERS = ["job_id", "project_name", "course_name", "content_area", "artifact_type", "file_path"]
DUPLICATE_HEADERS = ["job_id", "fingerprint", "policy", "duplicate_mode", "matched_job_id", "concept"]


def _ensure_sheet(wb, name: str, headers: list[str]):
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append(headers)
    elif [cell.value for cell in ws[1][: len(headers)]] != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
    return ws


def _load_or_create(path: Path):
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    for name, headers in {
        "Runs": RUNS_HEADERS,
        "Objectives": OBJECTIVE_HEADERS,
        "Cards": CARD_HEADERS,
        "Artifacts": ARTIFACT_HEADERS,
        "Duplicates": DUPLICATE_HEADERS,
    }.items():
        _ensure_sheet(wb, name, headers)
    return wb


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _append_unique(ws, row: list[Any], key_cols: list[int]):
    existing = set()
    for existing_row in ws.iter_rows(min_row=2, values_only=True):
        existing.add(tuple(existing_row[i] for i in key_cols))
    key = tuple(row[i] for i in key_cols)
    if key not in existing:
        ws.append(row)


def update_curriculum_master(state: dict[str, Any]) -> str:
    taxonomy = state["taxonomy"]["data"]
    deployment = state.get("deployment") or {}
    batch_dir = Path(deployment.get("batch_dir") or "")
    course_name = taxonomy.get("course_name") or taxonomy.get("subject_area") or "course"
    content_area = taxonomy.get("content_area") or "content"
    concept = taxonomy.get("concept") or "concept"
    project_name = state.get("input", {}).get("project_name") or taxonomy.get("project_name") or "curriculum_project"

    MASTER_ROOT.mkdir(parents=True, exist_ok=True)
    MASTER_CSV_DIR.mkdir(parents=True, exist_ok=True)
    workbook_path = MASTER_ROOT / workbook_filename(course_name, content_area, None, "curriculum_master")
    wb = _load_or_create(workbook_path)
    ws_runs = wb["Runs"]
    ws_obj = wb["Objectives"]
    ws_cards = wb["Cards"]
    ws_art = wb["Artifacts"]
    ws_dup = wb["Duplicates"]

    duplicate_check = state.get("duplicate_check") or {}
    _append_unique(
        ws_runs,
        [
            state.get("job_id"), project_name, course_name, content_area, taxonomy.get("subject_area"), taxonomy.get("specialty_area"),
            concept, state.get("input", {}).get("build_profile", "standard"), state.get("input", {}).get("deck_style", "clean_academic"),
            state.get("input", {}).get("duplicate_policy") or state.get("input", {}).get("duplicate_action") or taxonomy.get("duplicate_policy") or "create_new_version",
            state.get("duplicate_mode", "build"), (duplicate_check.get("matched") or {}).get("job_id", ""), state.get("status"),
            str(batch_dir), str(workbook_path),
        ],
        [0],
    )

    for row in _read_csv_rows(batch_dir / "curriculum_objectives.csv"):
        _append_unique(
            ws_obj,
            [
                state.get("job_id"), project_name, course_name, content_area,
                row.get("objective_id"), row.get("concept"), row.get("exemplar"), row.get("nclex_category"),
                row.get("ncjmm_primary"), row.get("priority_framework"), row.get("description"),
            ],
            [4],
        )

    for row in _read_csv_rows(batch_dir / "knowledge_cards.csv"):
        _append_unique(
            ws_cards,
            [
                state.get("job_id"), project_name, course_name, content_area,
                row.get("card_id"), row.get("card_title"), row.get("concept"), row.get("subject_area"),
                row.get("specialty_area"), row.get("status"), row.get("evidence_status"),
            ],
            [4],
        )

    build_artifacts = (state.get("build") or {}).get("artifacts", {})
    for artifact_type, path in build_artifacts.items():
        _append_unique(ws_art, [state.get("job_id"), project_name, course_name, content_area, artifact_type, path], [0, 4, 5])
    for csv_file in ["knowledge_cards.csv", "curriculum_objectives.csv", "card_objective_map.csv", "card_sources.csv", "card_links.csv"]:
        path = batch_dir / csv_file
        if path.exists():
            _append_unique(ws_art, [state.get("job_id"), project_name, course_name, content_area, csv_file, str(path)], [0, 4, 5])

    if duplicate_check:
        _append_unique(
            ws_dup,
            [
                state.get("job_id"),
                duplicate_check.get("fingerprint", ""),
                state.get("input", {}).get("duplicate_policy") or state.get("input", {}).get("duplicate_action") or taxonomy.get("duplicate_policy") or "create_new_version",
                state.get("duplicate_mode", "build"),
                (duplicate_check.get("matched") or {}).get("job_id", ""),
                concept,
            ],
            [0],
        )

    wb.save(workbook_path)
    MASTER_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(MASTER_XLSX)
    return str(workbook_path)
