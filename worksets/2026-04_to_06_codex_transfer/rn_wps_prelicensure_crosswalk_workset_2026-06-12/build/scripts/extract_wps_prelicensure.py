from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\RHarrity\Documents\Codex\rn_wps_prelicensure_crosswalk_workset_2026-06-12")
SOURCE_WORKBOOK = Path(r"C:\Users\RHarrity\Documents\Prelicensure Workbook.xlsx")
WPS_ROOT = Path(r"C:\Users\RHarrity\Downloads\Registered Nurse training Department of Education")
INTERMEDIATE_DIR = ROOT / "build" / "intermediate"
DATA_PATH = INTERMEDIATE_DIR / "rn_wps_prelicensure_crosswalk_data.json"
MANIFEST_PATH = ROOT / "workset_manifest.json"

W_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
SUBCOMPETENCY_RE = re.compile(r"^\d{1,2}\.\d+[a-z]$")
COMPETENCY_RE = re.compile(r"^(\d{1,2}\.\d+)\s+(.+)$")
AACN_DOMAIN_RE = re.compile(r"Domain\s+(\d{1,2})\s*:?\s*(.*)$", re.IGNORECASE)

STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "based", "be", "by", "care", "for", "from",
    "in", "into", "is", "may", "of", "on", "or", "patient", "patients", "the", "their",
    "to", "with", "within", "nurse", "nurses", "nursing",
}
ACTION_VERBS = {
    "administer", "advise", "assess", "collaborate", "communicate", "coordinate", "develop",
    "document", "educate", "evaluate", "identify", "implement", "inform", "maintain",
    "manage", "monitor", "plan", "prevent", "promote", "provide", "record", "report",
    "support", "teach", "treat",
}
MATCH_THRESHOLD = 0.30
AMBIGUITY_DELTA = 0.04


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return "sha256:" + h.hexdigest()


def source_record(path: Path, source_kind: str, canonical_input: bool) -> dict[str, Any]:
    stat = path.stat()
    return {
        "source_id": f"SRC_{len(source_records) + 1:03d}",
        "source_kind": source_kind,
        "canonical_input": canonical_input,
        "source_name": path.name,
        "source_path": str(path),
        "source_hash": sha256_file(path),
        "source_size_bytes": stat.st_size,
        "source_modified": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
    }


def slug(value: str, prefix: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return f"{prefix}_{base or 'blank'}"


def normalize_text(value: str) -> str:
    value = value.lower().replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def tokens(value: str) -> set[str]:
    out = set()
    for token in normalize_text(value).split():
        if len(token) < 3 or token in STOPWORDS:
            continue
        out.add(token)
    return out


def strip_task_label(value: str) -> str:
    return re.sub(r"^[A-Z]\.\s+", "", value).strip()


def parse_aacn_domain(raw: str) -> dict[str, str]:
    match = AACN_DOMAIN_RE.search(raw or "")
    if not match:
        return {
            "aacn_domain_id": "AACN_DOMAIN_UNKNOWN",
            "aacn_domain_number": "",
            "aacn_domain_name": raw,
            "aacn_domain_raw": raw,
        }
    number = match.group(1)
    name = match.group(2).strip()
    return {
        "aacn_domain_id": f"AACN_DOMAIN_{number.zfill(2)}",
        "aacn_domain_number": number,
        "aacn_domain_name": name,
        "aacn_domain_raw": raw,
    }


def parse_competency(raw: str) -> tuple[str, str]:
    match = COMPETENCY_RE.match(raw or "")
    if not match:
        return "", raw
    return match.group(1), match.group(2).strip()


def split_codes(raw: str) -> list[str]:
    normalized = re.sub(r"\.\s+(?=\d{1,2}\.\d+[a-z]\b)", ", ", raw)
    parts = re.split(r"[,;\n]+", normalized)
    return [p.strip() for p in parts if p.strip()]


def cell(ws, row: int, col: int) -> str:
    return clean(ws.cell(row=row, column=col).value)


def row_has_any_value(ws, row: int) -> bool:
    return any(cell(ws, row, col) for col in range(1, ws.max_column + 1))


def is_source_footer_row(statement: str, row: int) -> bool:
    if row < 214:
        return False
    if statement.startswith("http://") or statement.startswith("https://"):
        return True
    markers = ["PhD", "DNP", "RN", "CNE", "FAAN", "University", "School of Nursing"]
    return "/" in statement and any(marker in statement for marker in markers)


def collect_domain_starts(ws) -> list[tuple[int, dict[str, str]]]:
    starts = []
    for col in range(1, ws.max_column + 1):
        raw = cell(ws, 2, col)
        if "Domain" in raw and "AACN" in raw:
            starts.append((col, parse_aacn_domain(raw)))
    return starts


def domain_for_col(domain_starts: list[tuple[int, dict[str, str]]], col: int) -> dict[str, str]:
    selected = None
    for start_col, domain in domain_starts:
        if start_col <= col:
            selected = domain
        else:
            break
    return selected or {
        "aacn_domain_id": "AACN_DOMAIN_UNKNOWN",
        "aacn_domain_number": "",
        "aacn_domain_name": "",
        "aacn_domain_raw": "",
    }


def collect_competency_columns(ws) -> dict[int, dict[str, Any]]:
    domain_starts = collect_domain_starts(ws)
    columns = {}
    for col in range(4, ws.max_column + 1):
        raw = cell(ws, 4, col)
        has_values = any(cell(ws, row, col) for row in range(6, ws.max_row + 1))
        if not raw and has_values and cell(ws, 5, col).lower() == "subcompetencies":
            domain = domain_for_col(domain_starts, col)
            columns[col] = {
                "source_col": get_column_letter(col),
                "source_col_index": col,
                "source_header_cell": f"{get_column_letter(col)}4",
                "competency_id": "",
                "competency_name": "",
                "competency_raw": "",
                "missing_competency_header": True,
                **domain,
            }
            continue
        if not raw:
            continue
        competency_id, competency_name = parse_competency(raw)
        if not competency_id:
            continue
        domain = domain_for_col(domain_starts, col)
        columns[col] = {
            "source_col": get_column_letter(col),
            "source_col_index": col,
            "source_header_cell": f"{get_column_letter(col)}4",
            "competency_id": competency_id,
            "competency_name": competency_name,
            "competency_raw": raw,
            "missing_competency_header": False,
            **domain,
        }
    return columns


def parse_prelicensure() -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(SOURCE_WORKBOOK, data_only=False)
    inventory_sheets = []
    for ws in wb.worksheets:
        competency_columns = collect_competency_columns(ws)
        data_rows = 0
        mapped_cells = 0
        for row in range(6, ws.max_row + 1):
            if row_has_any_value(ws, row):
                data_rows += 1
            for col in competency_columns:
                if cell(ws, row, col):
                    mapped_cells += 1
        inventory_sheets.append(
            {
                "sheet_name": ws.title,
                "dimension": ws.calculate_dimension(),
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "data_rows_from_6": data_rows,
                "competency_column_count": len(competency_columns),
                "mapped_cells_pre_split": mapped_cells,
            }
        )

    entry = wb["Entry Level"]
    entry_comp_columns = collect_competency_columns(entry)
    facts = []
    exclusions = []
    current_domain_raw = ""
    current_ksa_raw = ""
    qsen_sequence = 0
    for row in range(6, entry.max_row + 1):
        domain_value = cell(entry, row, 1)
        ksa_value = cell(entry, row, 2)
        statement = cell(entry, row, 3)
        if domain_value:
            current_domain_raw = domain_value
        if ksa_value:
            current_ksa_raw = ksa_value
        if not statement and not row_has_any_value(entry, row):
            continue
        if not statement:
            continue
        if is_source_footer_row(statement, row):
            exclusions.append(
                {
                    "source_workbook": SOURCE_WORKBOOK.name,
                    "source_sheet": "Entry Level",
                    "source_row": row,
                    "source_cell": f"C{row}",
                    "exclusion_reason": "source_footer_or_reference_row",
                    "raw_value": statement,
                }
            )
            continue
        qsen_sequence += 1
        codes = []
        raw_cells = []
        for col, meta in entry_comp_columns.items():
            raw = cell(entry, row, col)
            if not raw:
                continue
            raw_cells.append(f"{get_column_letter(col)}{row}: {raw}")
            for code in split_codes(raw):
                codes.append(code)
        qsen_domain_name = current_domain_raw.split(":", 1)[0].strip() if ":" in current_domain_raw else current_domain_raw
        fact = {
            "qsen_statement_id": f"QSEN_{qsen_sequence:04d}",
            "source_workbook": SOURCE_WORKBOOK.name,
            "source_sheet": "Entry Level",
            "source_row": row,
            "qsen_domain_raw": current_domain_raw,
            "qsen_domain_name": qsen_domain_name,
            "ksa_raw": current_ksa_raw,
            "qsen_statement_raw": statement,
            "qsen_statement_normalized": normalize_text(statement),
            "aacn_codes_raw": "; ".join(raw_cells),
            "aacn_subcompetency_codes": "; ".join(codes),
            "aacn_subcompetency_count": len(codes),
            "malformed_aacn_code_count": sum(1 for code in codes if not SUBCOMPETENCY_RE.match(code)),
        }
        facts.append(fact)

    inventory = {
        "source_path": str(SOURCE_WORKBOOK),
        "source_name": SOURCE_WORKBOOK.name,
        "source_hash": sha256_file(SOURCE_WORKBOOK),
        "source_size_bytes": SOURCE_WORKBOOK.stat().st_size,
        "source_modified": datetime.fromtimestamp(SOURCE_WORKBOOK.stat().st_mtime).isoformat(timespec="seconds"),
        "sheet_count": len(wb.worksheets),
        "sheets": inventory_sheets,
        "entry_level_present": "Entry Level" in wb.sheetnames,
    }
    return inventory, facts, exclusions


def docx_cell_text(tc: ET.Element) -> str:
    parts = []
    for p in tc.findall(".//w:p", W_NS):
        text = "".join(t.text or "" for t in p.findall(".//w:t", W_NS)).strip()
        if text:
            parts.append(re.sub(r"\s+", " ", text))
    return " | ".join(parts).strip()


def parse_prefixed(value: str, prefix: str) -> str:
    for part in value.split("|"):
        part = part.strip()
        if part.lower().startswith(prefix.lower()):
            return part.split(":", 1)[1].strip() if ":" in part else part
    return ""


def infer_variant(filename: str, apprenticeship_type: str) -> str:
    lower = filename.lower()
    if "_competency_" in lower:
        return "competency_based"
    if "_hybrid_" in lower:
        return "hybrid"
    if "_time_" in lower:
        return "time_based"
    if "Hybrid" in apprenticeship_type:
        return "hybrid"
    if "Competency" in apprenticeship_type:
        return "competency_based"
    return "time_based"


def parse_wps_docx(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    with zipfile.ZipFile(path) as z:
        xml = ET.fromstring(z.read("word/document.xml"))
    tables = xml.findall(".//w:tbl", W_NS)
    rows_by_table = []
    for tbl in tables:
        table_rows = []
        for tr in tbl.findall("./w:tr", W_NS):
            cells = [docx_cell_text(tc) for tc in tr.findall("./w:tc", W_NS)]
            if any(cells):
                table_rows.append(cells)
        rows_by_table.append(table_rows)

    meta_rows = rows_by_table[0]
    role_title = meta_rows[0][0] if meta_rows and meta_rows[0] else ""
    job_description = meta_rows[1][0] if len(meta_rows) > 1 and meta_rows[1] else ""
    rapids_code = parse_prefixed(" | ".join(meta_rows[2]) if len(meta_rows) > 2 else "", "RAPIDS Code")
    onet_code = parse_prefixed(" | ".join(meta_rows[2]) if len(meta_rows) > 2 else "", "O*NET Code")
    program_length = parse_prefixed(meta_rows[3][0] if len(meta_rows) > 3 and meta_rows[3] else "", "Estimated Program Length")
    apprenticeship_type_raw = parse_prefixed(meta_rows[4][0] if len(meta_rows) > 4 and meta_rows[4] else "", "Apprenticeship Type")
    variant = infer_variant(path.name, apprenticeship_type_raw)
    rapids_base = re.sub(r"\D+$", "", rapids_code)
    source_file_id = f"WPSFILE_{rapids_code or path.stem}"

    metadata = {
        "wps_file_id": source_file_id,
        "source_file": path.name,
        "source_path": str(path),
        "table_count": len(tables),
        "role_title_raw": role_title,
        "job_description_raw": job_description,
        "rapids_code": rapids_code,
        "rapids_base": rapids_base,
        "onet_code": onet_code,
        "estimated_program_length_raw": program_length,
        "apprenticeship_type_raw": apprenticeship_type_raw,
        "variant": variant,
    }

    processes = []
    tasks = []
    for table_index, table_rows in enumerate(rows_by_table[1:], start=2):
        if not table_rows:
            continue
        process_raw = table_rows[0][0].strip()
        if not process_raw:
            continue
        process_id = slug(process_raw, "WPS_PROCESS")
        processes.append(
            {
                "wps_process_id": process_id,
                "process_raw": process_raw,
                "process_normalized": normalize_text(process_raw),
                "source_file": path.name,
                "rapids_code": rapids_code,
                "variant": variant,
                "table_index": table_index,
            }
        )
        for row_offset, cells in enumerate(table_rows[1:], start=2):
            first = cells[0].strip() if cells else ""
            if not first or first.lower().startswith("competencies"):
                continue
            if first.lower().startswith("total hours"):
                continue
            label_match = re.match(r"^([A-Z])\.\s+", first)
            task_label = label_match.group(1) if label_match else ""
            task_text = strip_task_label(first)
            task_id = f"WPS_{rapids_code}_{variant.upper()}_T{table_index:02d}_R{row_offset:02d}_{task_label or 'X'}"
            data_cells_raw = " | ".join(cells[1:]).strip()
            tasks.append(
                {
                    "wps_task_id": task_id,
                    "wps_file_id": source_file_id,
                    "source_file": path.name,
                    "source_path": str(path),
                    "rapids_code": rapids_code,
                    "rapids_base": rapids_base,
                    "role_title_raw": role_title,
                    "variant": variant,
                    "apprenticeship_type_raw": apprenticeship_type_raw,
                    "wps_process_id": process_id,
                    "process_raw": process_raw,
                    "task_label": task_label,
                    "task_raw": first,
                    "task_text_raw": task_text,
                    "task_normalized": normalize_text(task_text),
                    "source_table_index": table_index,
                    "source_row_index": row_offset,
                    "data_cells_raw": data_cells_raw,
                }
            )
    return metadata, processes, tasks


def similarity(task: dict[str, Any], fact: dict[str, Any]) -> float:
    task_text = task["task_text_raw"]
    qsen_text = fact["qsen_statement_raw"]
    task_norm = normalize_text(task_text)
    qsen_norm = fact["qsen_statement_normalized"]
    if not task_norm or not qsen_norm:
        return 0.0
    if task_norm == qsen_norm:
        return 1.0

    task_tokens = tokens(task_text)
    qsen_tokens = tokens(qsen_text + " " + fact.get("ksa_raw", "") + " " + fact.get("qsen_domain_raw", ""))
    if not task_tokens or not qsen_tokens:
        return 0.0
    overlap = task_tokens & qsen_tokens
    union = task_tokens | qsen_tokens
    jaccard = len(overlap) / len(union)
    coverage = len(overlap) / len(task_tokens)
    containment = 0.0
    if task_norm in qsen_norm or qsen_norm in task_norm:
        containment = 0.35
    verb_bonus = 0.05 if (task_tokens & ACTION_VERBS & qsen_tokens) else 0.0
    phrase_bonus = 0.05 if len(overlap) >= 4 else 0.0
    score = (0.55 * coverage) + (0.35 * jaccard) + containment + verb_bonus + phrase_bonus
    return round(min(score, 1.0), 4)


def build_bridge(wps_tasks: list[dict[str, Any]], qsen_facts: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    duplicate_contexts = defaultdict(set)
    for task in wps_tasks:
        duplicate_contexts[task["task_normalized"]].add(normalize_text(task["process_raw"]))

    bridges = []
    review_queue = []
    for idx, task in enumerate(wps_tasks, start=1):
        scored = sorted(
            ((similarity(task, fact), fact) for fact in qsen_facts),
            key=lambda item: item[0],
            reverse=True,
        )
        best_score, best_fact = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0
        reasons = []
        if best_score < MATCH_THRESHOLD:
            reasons.append("no_confident_match")
        if best_score >= MATCH_THRESHOLD and (best_score - second_score) <= AMBIGUITY_DELTA:
            reasons.append("ambiguous_match")
        if task["task_normalized"] and len(duplicate_contexts[task["task_normalized"]]) > 1:
            reasons.append("duplicate_task_different_process")
        needs_review = bool(reasons)
        qsen_fact = best_fact if best_score > 0 else {}
        bridge = {
            "bridge_id": f"WPS_QSEN_BRIDGE_{idx:04d}",
            "wps_task_id": task["wps_task_id"],
            "source_file": task["source_file"],
            "rapids_code": task["rapids_code"],
            "rapids_base": task["rapids_base"],
            "variant": task["variant"],
            "process_raw": task["process_raw"],
            "task_raw": task["task_raw"],
            "task_text_raw": task["task_text_raw"],
            "proposed_qsen_statement_id": qsen_fact.get("qsen_statement_id", ""),
            "proposed_qsen_domain_name": qsen_fact.get("qsen_domain_name", ""),
            "proposed_ksa_raw": qsen_fact.get("ksa_raw", ""),
            "proposed_qsen_statement_raw": qsen_fact.get("qsen_statement_raw", ""),
            "proposed_aacn_subcompetency_codes": qsen_fact.get("aacn_subcompetency_codes", ""),
            "match_score": best_score,
            "second_best_score": second_score,
            "needs_review": needs_review,
            "review_reason": "; ".join(reasons),
            "mapping_method": "local_lexical_rule_based",
        }
        bridges.append(bridge)
        if needs_review:
            review_queue.append(
                {
                    "review_id": f"WPS_REVIEW_{len(review_queue) + 1:04d}",
                    "record_type": "suggested_wps_prelicensure_bridge",
                    "bridge_id": bridge["bridge_id"],
                    "wps_task_id": task["wps_task_id"],
                    "source_file": task["source_file"],
                    "rapids_code": task["rapids_code"],
                    "variant": task["variant"],
                    "process_raw": task["process_raw"],
                    "task_raw": task["task_raw"],
                    "proposed_qsen_statement_id": bridge["proposed_qsen_statement_id"],
                    "match_score": best_score,
                    "review_reason": bridge["review_reason"],
                }
            )
    return bridges, review_queue


def build_variant_delta(wps_tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    variants = [
        ("2074", "time_based"),
        ("2074", "competency_based"),
        ("2074", "hybrid"),
        ("3056", "time_based"),
        ("3056", "competency_based"),
        ("3056", "hybrid"),
    ]
    grouped = {}
    for task in wps_tasks:
        key = (task["process_normalized"] if "process_normalized" in task else normalize_text(task["process_raw"]), task["task_normalized"])
        grouped.setdefault(
            key,
            {
                "variant_delta_id": f"WPS_DELTA_{len(grouped) + 1:04d}",
                "process_raw": task["process_raw"],
                "task_text_raw": task["task_text_raw"],
                "normalized_key": " | ".join(key),
                "source_occurrence_count": 0,
            },
        )
        grouped[key]["source_occurrence_count"] += 1
        grouped[key][f"present_{task['rapids_base']}_{task['variant']}"] = True
    rows = []
    for row in grouped.values():
        present = 0
        for rapids, variant in variants:
            col = f"present_{rapids}_{variant}"
            value = bool(row.get(col, False))
            row[col] = value
            present += 1 if value else 0
        row["variant_presence_count"] = present
        row["variant_delta_status"] = "present_all_six" if present == 6 else "variant_difference"
        rows.append(row)
    return sorted(rows, key=lambda r: (r["process_raw"], r["task_text_raw"]))


def build_coverage_summary(wps_tasks: list[dict[str, Any]], bridges: list[dict[str, Any]], qsen_facts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    by_variant = defaultdict(lambda: {"task_count": 0, "review_count": 0, "confident_count": 0})
    for bridge in bridges:
        key = f"{bridge['rapids_code']} / {bridge['variant']}"
        by_variant[key]["task_count"] += 1
        if bridge["needs_review"]:
            by_variant[key]["review_count"] += 1
        else:
            by_variant[key]["confident_count"] += 1
    for label, counts in sorted(by_variant.items()):
        rows.append({"coverage_type": "wps_variant", "label": label, **counts})

    by_process = defaultdict(lambda: {"task_count": 0, "review_count": 0, "confident_count": 0})
    for bridge in bridges:
        label = bridge["process_raw"]
        by_process[label]["task_count"] += 1
        if bridge["needs_review"]:
            by_process[label]["review_count"] += 1
        else:
            by_process[label]["confident_count"] += 1
    for label, counts in sorted(by_process.items()):
        rows.append({"coverage_type": "wps_process", "label": label, **counts})

    matched_qsen = {b["proposed_qsen_statement_id"] for b in bridges if b["proposed_qsen_statement_id"]}
    rows.append(
        {
            "coverage_type": "prelicensure",
            "label": "QSEN statements matched by at least one WPS task",
            "task_count": len(matched_qsen),
            "review_count": "",
            "confident_count": "",
        }
    )
    rows.append(
        {
            "coverage_type": "prelicensure",
            "label": "Total canonical QSEN statements",
            "task_count": len(qsen_facts),
            "review_count": "",
            "confident_count": "",
        }
    )
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


source_records: list[dict[str, Any]] = []


def main() -> None:
    INTERMEDIATE_DIR.mkdir(parents=True, exist_ok=True)
    source_records.append(source_record(SOURCE_WORKBOOK, "prelicensure_workbook", True))
    for path in sorted(WPS_ROOT.glob("wps_*.docx")):
        source_records.append(source_record(path, "wps_docx", True))
    for path in sorted(WPS_ROOT.glob("*.wbk")):
        source_records.append(source_record(path, "word_backup", False))

    prelic_inventory, qsen_facts, prelic_exclusions = parse_prelicensure()

    wps_metadata = []
    wps_process_observations = []
    wps_tasks = []
    for path in sorted(WPS_ROOT.glob("wps_*.docx")):
        metadata, processes, tasks = parse_wps_docx(path)
        wps_metadata.append(metadata)
        wps_process_observations.extend(processes)
        wps_tasks.extend(tasks)

    process_rollup = {}
    for proc in wps_process_observations:
        key = proc["wps_process_id"]
        process_rollup.setdefault(
            key,
            {
                "wps_process_id": key,
                "process_raw": proc["process_raw"],
                "process_normalized": proc["process_normalized"],
                "source_file_count": 0,
                "source_files": set(),
                "task_count": 0,
            },
        )
        process_rollup[key]["source_files"].add(proc["source_file"])
    for task in wps_tasks:
        process_rollup[task["wps_process_id"]]["task_count"] += 1
    dim_wps_process = []
    for row in process_rollup.values():
        source_files = sorted(row.pop("source_files"))
        row["source_file_count"] = len(source_files)
        row["source_files"] = "; ".join(source_files)
        dim_wps_process.append(row)
    dim_wps_process.sort(key=lambda r: r["process_raw"])

    bridges, review_queue = build_bridge(wps_tasks, qsen_facts)
    variant_delta = build_variant_delta(wps_tasks)
    coverage_summary = build_coverage_summary(wps_tasks, bridges, qsen_facts)

    source_hash_after = {record["source_path"]: sha256_file(Path(record["source_path"])) for record in source_records}
    source_files_modified = any(record["source_hash"] != source_hash_after[record["source_path"]] for record in source_records)

    docx_table_counts = {row["source_file"]: row["table_count"] for row in wps_metadata}
    rapids_codes = sorted(row["rapids_code"] for row in wps_metadata)
    data = {
        "metadata": {
            "workset_id": "rn_wps_prelicensure_crosswalk_workset_2026-06-12",
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "mapping_policy": "local_only_conservative_lexical_rule_based",
            "match_threshold": MATCH_THRESHOLD,
            "ambiguity_delta": AMBIGUITY_DELTA,
            "source_files_modified": source_files_modified,
        },
        "prelicensure_inventory": prelic_inventory,
        "wps_source_files": source_records,
        "wps_metadata": wps_metadata,
        "dim_wps_process": dim_wps_process,
        "fact_wps_tasks": wps_tasks,
        "fact_prelicensure_qsen": qsen_facts,
        "suggested_wps_prelicensure_bridge": bridges,
        "variant_delta": variant_delta,
        "coverage_summary": coverage_summary,
        "review_queue": review_queue,
        "source_exclusions": prelic_exclusions,
        "reconciliation": {
            "prelicensure_sheet_count_is_11": prelic_inventory["sheet_count"] == 11,
            "entry_level_present": prelic_inventory["entry_level_present"],
            "wps_docx_file_count_is_6": len(wps_metadata) == 6,
            "all_wps_docx_have_40_tables": all(count == 40 for count in docx_table_counts.values()),
            "required_rapids_codes_present": rapids_codes == ["2074", "2074CB", "2074HY", "3056", "3056CB", "3056HY"],
            "all_wps_tasks_have_source_coordinates": all(
                task["source_file"] and task["source_table_index"] and task["source_row_index"] and task["task_raw"]
                for task in wps_tasks
            ),
            "source_files_modified": source_files_modified,
        },
    }

    DATA_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    write_csv(INTERMEDIATE_DIR / "fact_wps_tasks.csv", wps_tasks)
    write_csv(INTERMEDIATE_DIR / "fact_prelicensure_qsen.csv", qsen_facts)
    write_csv(INTERMEDIATE_DIR / "suggested_wps_prelicensure_bridge.csv", bridges)
    write_csv(INTERMEDIATE_DIR / "review_queue.csv", review_queue)

    manifest = {
        "workset_id": data["metadata"]["workset_id"],
        "created_at": data["metadata"]["generated_at"],
        "source_policy": "read_only_sources",
        "mapping_policy": data["metadata"]["mapping_policy"],
        "canonical_inputs": [
            record for record in source_records if record["canonical_input"]
        ],
        "provenance_only_inputs": [
            record for record in source_records if not record["canonical_input"]
        ],
        "outputs_planned": [
            str(ROOT / "outputs" / "rn_wps_prelicensure_crosswalk_release.xlsx"),
            str(ROOT / "rn_wps_prelicensure_crosswalk_qa.json"),
        ],
        "notes": [
            "WBK backup inventoried as provenance only.",
            "No OpenAI or external semantic API calls are used for mapping.",
            "Requested sheet name Suggested_WPS_Prelicensure_Bridge is shortened in Excel output to Suggested_WPS_Prelic_Bridge due to Excel's 31-character sheet-name limit.",
        ],
    }
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "data_path": str(DATA_PATH),
                "manifest_path": str(MANIFEST_PATH),
                "wps_docx_files": len(wps_metadata),
                "wps_tasks": len(wps_tasks),
                "qsen_facts": len(qsen_facts),
                "suggested_bridge_rows": len(bridges),
                "review_queue_rows": len(review_queue),
                "reconciliation": data["reconciliation"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
