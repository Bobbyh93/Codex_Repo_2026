from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\RHarrity\Documents\Codex\prelicensure_crosswalk_workset_2026-06-11")
DATA_PATH = ROOT / "build" / "intermediate" / "crosswalk_data.json"
SOURCE_PATH = Path(r"C:\Users\RHarrity\Downloads\Prelicensure Workbook.xlsx")
OUTPUT_DIR = ROOT / "outputs"
MD_PATH = OUTPUT_DIR / "remaining_review_packet.md"
CSV_PATH = OUTPUT_DIR / "remaining_review_queue.csv"
JSON_PATH = OUTPUT_DIR / "remaining_review_context.json"


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    facts = {row["qsen_statement_id"]: row for row in data["fact_qsen_statements"]}
    bridges = {row["bridge_id"]: row for row in data["bridge_qsen_aacn"]}
    reviews = data["review_queue"]

    workbook = load_workbook(SOURCE_PATH, read_only=False, data_only=False)
    sheet = workbook["Entry Level"]
    header_context = {
        "D2": sheet["D2"].value,
        "D3": sheet["D3"].value,
        "D4": sheet["D4"].value,
        "D5": sheet["D5"].value,
        "E4": sheet["E4"].value,
        "F4": sheet["F4"].value,
        "merged_ranges_touching_d_rows_2_5": [
            str(rng)
            for rng in sheet.merged_cells.ranges
            if rng.min_col <= 4 <= rng.max_col and rng.min_row <= 5 and rng.max_row >= 2
        ],
    }

    rows = []
    for review in reviews:
        bridge = bridges.get(review.get("bridge_id"), {})
        fact = facts.get(review["qsen_statement_id"], {})
        rows.append(
            {
                "review_id": review["review_id"],
                "qsen_statement_id": review["qsen_statement_id"],
                "bridge_id": review.get("bridge_id", ""),
                "source_sheet": review["source_sheet"],
                "source_row": review["source_row"],
                "source_cell": review["source_cell"],
                "raw_mapping_value": review["raw_value"],
                "subcompetency_code": bridge.get("subcompetency_code", review["raw_value"]),
                "implied_parent_competency_code": str(bridge.get("subcompetency_code", review["raw_value"]))[:-1],
                "qsen_domain_name": fact.get("qsen_domain_name", ""),
                "ksa": fact.get("ksa_raw", ""),
                "qsen_statement_raw": fact.get("qsen_statement_raw", ""),
                "review_reason": review["review_reason"],
                "recommended_action": "Confirm the missing AACN 1.1 competency title from an authoritative AACN source before clearing review.",
            }
        )

    fieldnames = [
        "review_id",
        "qsen_statement_id",
        "bridge_id",
        "source_sheet",
        "source_row",
        "source_cell",
        "raw_mapping_value",
        "subcompetency_code",
        "implied_parent_competency_code",
        "qsen_domain_name",
        "ksa",
        "qsen_statement_raw",
        "review_reason",
        "recommended_action",
    ]
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    JSON_PATH.write_text(
        json.dumps(
            {
                "source_workbook": str(SOURCE_PATH),
                "remaining_review_count": len(rows),
                "review_reason_counts": data.get("metadata", {}),
                "header_context": header_context,
                "rows": rows,
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    lines = [
        "# Remaining Review Packet",
        "",
        "Purpose: isolate the final unresolved rows from the reviewed prelicensure QSEN-to-AACN crosswalk release.",
        "",
        "## Status",
        "",
        f"- Remaining review rows: {len(rows)}",
        "- Remaining reason: `missing_competency_header`",
        "- Source files were not modified.",
        "- These rows are preserved in the reviewed release workbook and should not be cleared until the missing AACN 1.1 competency title is confirmed from an authoritative source.",
        "",
        "## Source Header Evidence",
        "",
        "| Cell | Value |",
        "|---|---|",
    ]
    for cell_ref in ["D2", "D3", "D4", "D5", "E4", "F4"]:
        value = header_context[cell_ref]
        display = "" if value is None else str(value).replace("\n", " ")
        lines.append(f"| {cell_ref} | {display} |")
    lines.extend(
        [
            "",
            "Merged ranges touching column D rows 2-5:",
            "",
            *[f"- `{rng}`" for rng in header_context["merged_ranges_touching_d_rows_2_5"]],
            "",
            "Interpretation: column D is inside the Domain 1 subcompetency region, but `D4` has no parent competency title. The source cells map to `1.1b` and `1.1d`, which imply parent competency `1.1`; the workbook itself does not provide the `1.1` title in row 4.",
            "",
            "## Remaining Rows",
            "",
            "| Review ID | Source Cell | Code | Implied Parent | KSA | QSEN Statement |",
            "|---|---|---|---|---|---|",
        ]
    )
    for row in rows:
        statement = row["qsen_statement_raw"].replace("\n", " ")
        if len(statement) > 180:
            statement = statement[:177] + "..."
        lines.append(
            f"| {row['review_id']} | {row['source_cell']} | {row['subcompetency_code']} | {row['implied_parent_competency_code']} | {row['ksa']} | {statement} |"
        )
    lines.extend(
        [
            "",
            "## Recommended Resolution",
            "",
            "1. Confirm the official AACN Entry Level competency `1.1` title from an authoritative AACN source.",
            "2. If confirmed, update the release process to populate the parent competency metadata for column D while preserving the workbook's blank `D4` in provenance fields.",
            "3. Rebuild the release workbook and QA manifest; the review queue should then clear if no other source issues appear.",
            "",
            "## Files",
            "",
            f"- CSV review queue: `{CSV_PATH}`",
            f"- JSON context: `{JSON_PATH}`",
            "",
        ]
    )
    MD_PATH.write_text("\n".join(lines), encoding="utf-8")

    print(
        json.dumps(
            {
                "remaining_review_packet": str(MD_PATH),
                "remaining_review_csv": str(CSV_PATH),
                "remaining_review_context": str(JSON_PATH),
                "remaining_review_count": len(rows),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
