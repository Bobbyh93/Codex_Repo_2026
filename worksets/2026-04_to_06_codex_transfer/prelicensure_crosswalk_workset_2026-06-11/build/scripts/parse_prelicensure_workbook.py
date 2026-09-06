from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
SOURCE_PATH = Path(r"C:\Users\RHarrity\Downloads\Prelicensure Workbook.xlsx")
INTERMEDIATE_DIR = ROOT / "build" / "intermediate"
SOURCE_INVENTORY_PATH = INTERMEDIATE_DIR / "source_inventory.json"
CROSSWALK_DATA_PATH = INTERMEDIATE_DIR / "crosswalk_data.json"

SUBCOMPETENCY_RE = re.compile(r"^\d{1,2}\.\d+[a-z]$")
COMPETENCY_RE = re.compile(r"^(\d{1,2}\.\d+)\s+(.+)$")
AACN_DOMAIN_RE = re.compile(r"Domain\s+(\d{1,2})\s*:?\s*(.*)$", re.IGNORECASE)
SUPPLEMENTAL_COMPETENCY_TITLES = {
    "1.1": {
        "competency_name": "Demonstrate an understanding of the discipline of nursing's distinct perspective and where shared perspectives exist with other disciplines.",
        "competency_title_source": "AACN Domain 1: Knowledge for Nursing Practice",
        "competency_title_source_url": "https://www.aacnnursing.org/essentials/tool-kit/domains-concepts/knowledge-for-nursing-practice",
        "competency_title_resolution": "supplemented_from_official_aacn_domain_page",
    }
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return "sha256:" + h.hexdigest()


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def slug(value: str, prefix: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return f"{prefix}_{base or 'blank'}"


def split_codes(raw: str) -> list[str]:
    normalized = re.sub(r"\.\s+(?=\d{1,2}\.\d+[a-z]\b)", ", ", raw)
    parts = re.split(r"[,;\n]+", normalized)
    return [p.strip() for p in parts if p.strip()]


def is_source_footer_row(statement: str, row: int) -> bool:
    if row < 214:
        return False
    if statement.startswith("http://") or statement.startswith("https://"):
        return True
    credential_markers = ["PhD", "DNP", "RN", "CNE", "FAAN", "University", "School of Nursing"]
    return "/" in statement and any(marker in statement for marker in credential_markers)


def is_declared_no_alignment(statement: str) -> bool:
    return "No alignment identified" in statement


def parse_aacn_domain(raw: str) -> dict:
    match = AACN_DOMAIN_RE.search(raw or "")
    if not match:
        return {
            "aacn_domain_id": "AACN_DOMAIN_UNKNOWN",
            "aacn_domain_number": "",
            "aacn_domain_name": raw,
            "aacn_domain_raw": raw,
        }
    number = match.group(1)
    name = match.group(2).strip()
    return {
        "aacn_domain_id": f"AACN_DOMAIN_{number.zfill(2)}",
        "aacn_domain_number": number,
        "aacn_domain_name": name,
        "aacn_domain_raw": raw,
    }


def parse_competency(raw: str) -> tuple[str, str]:
    match = COMPETENCY_RE.match(raw or "")
    if not match:
        return "", raw
    return match.group(1), match.group(2).strip()


def cell(ws, row: int, col: int) -> str:
    return clean(ws.cell(row=row, column=col).value)


def row_has_any_value(ws, row: int) -> bool:
    return any(cell(ws, row, col) for col in range(1, ws.max_column + 1))


def collect_domain_starts(ws) -> list[tuple[int, dict]]:
    starts = []
    for col in range(1, ws.max_column + 1):
        raw = cell(ws, 2, col)
        if "Domain" in raw and "AACN" in raw:
            starts.append((col, parse_aacn_domain(raw)))
    return starts


def domain_for_col(domain_starts: list[tuple[int, dict]], col: int) -> dict:
    selected = None
    for start_col, domain in domain_starts:
        if start_col <= col:
            selected = domain
        else:
            break
    return selected or {
        "aacn_domain_id": "AACN_DOMAIN_UNKNOWN",
        "aacn_domain_number": "",
        "aacn_domain_name": "",
        "aacn_domain_raw": "",
    }


def collect_competency_columns(ws) -> dict[int, dict]:
    domain_starts = collect_domain_starts(ws)
    columns = {}
    for col in range(4, ws.max_column + 1):
        raw = cell(ws, 4, col)
        has_data_values = any(cell(ws, row, col) for row in range(6, ws.max_row + 1))
        if not raw and has_data_values and cell(ws, 5, col).lower() == "subcompetencies":
            domain = domain_for_col(domain_starts, col)
            columns[col] = {
                "source_sheet": ws.title,
                "source_col": get_column_letter(col),
                "source_col_index": col,
                "source_header_cell": f"{get_column_letter(col)}4",
                "competency_id": "",
                "competency_name": "",
                "competency_raw": "",
                "missing_competency_header": True,
                **domain,
            }
            continue
        if not raw:
            continue
        competency_id, competency_name = parse_competency(raw)
        if not competency_id:
            continue
        domain = domain_for_col(domain_starts, col)
        columns[col] = {
            "source_sheet": ws.title,
            "source_col": get_column_letter(col),
            "source_col_index": col,
            "source_header_cell": f"{get_column_letter(col)}4",
            "competency_id": competency_id,
            "competency_name": competency_name,
            "competency_raw": raw,
            "missing_competency_header": False,
            **domain,
        }
    return columns


def workbook_inventory(wb) -> dict:
    sheets = []
    for ws in wb.worksheets:
        data_rows = 0
        mapped_cells = 0
        competency_columns = collect_competency_columns(ws)
        for row in range(6, ws.max_row + 1):
            if row_has_any_value(ws, row):
                data_rows += 1
            for col in competency_columns:
                if cell(ws, row, col):
                    mapped_cells += 1
        sheets.append(
            {
                "sheet_name": ws.title,
                "dimension": ws.calculate_dimension(),
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "data_rows_from_6": data_rows,
                "competency_column_count": len(competency_columns),
                "mapped_cells_pre_split": mapped_cells,
            }
        )
    return {
        "source_path": str(SOURCE_PATH),
        "source_name": SOURCE_PATH.name,
        "source_hash": sha256_file(SOURCE_PATH),
        "source_size_bytes": SOURCE_PATH.stat().st_size,
        "source_modified": datetime.fromtimestamp(SOURCE_PATH.stat().st_mtime).isoformat(timespec="seconds"),
        "sheet_count": len(wb.worksheets),
        "sheets": sheets,
    }


def build_crosswalk_data(wb, inventory: dict) -> dict:
    entry = wb["Entry Level"]
    entry_comp_columns = collect_competency_columns(entry)

    facts = []
    bridges = []
    domains = {}
    ksas = {}
    aacn_domains = {}
    aacn_competencies = {}
    review_queue = []
    source_exclusions = []

    current_domain_raw = ""
    current_ksa_raw = ""
    qsen_sequence = 0

    for col_meta in entry_comp_columns.values():
        aacn_domains[col_meta["aacn_domain_id"]] = {
            "aacn_domain_id": col_meta["aacn_domain_id"],
            "aacn_domain_number": col_meta["aacn_domain_number"],
            "aacn_domain_name": col_meta["aacn_domain_name"],
            "aacn_domain_raw": col_meta["aacn_domain_raw"],
        }
        if col_meta["competency_id"]:
            aacn_competencies[col_meta["competency_id"]] = {
                "competency_id": col_meta["competency_id"],
                "competency_name": col_meta["competency_name"],
                "competency_raw": col_meta["competency_raw"],
                "aacn_domain_id": col_meta["aacn_domain_id"],
                "aacn_domain_number": col_meta["aacn_domain_number"],
                "source_sheet": col_meta["source_sheet"],
                "source_header_cell": col_meta["source_header_cell"],
                "competency_title_source": "source_workbook_header",
                "competency_title_source_url": "",
                "source_header_missing": False,
                "source_header_missing_resolved": False,
                "needs_review": False,
                "review_reason": "",
            }

    for row in range(6, entry.max_row + 1):
        domain_value = cell(entry, row, 1)
        ksa_value = cell(entry, row, 2)
        statement = cell(entry, row, 3)

        if domain_value:
            current_domain_raw = domain_value
        if ksa_value:
            current_ksa_raw = ksa_value
        if not statement and not row_has_any_value(entry, row):
            continue
        if not statement:
            continue
        if is_source_footer_row(statement, row):
            source_exclusions.append(
                {
                    "source_workbook": SOURCE_PATH.name,
                    "source_sheet": "Entry Level",
                    "source_row": row,
                    "source_cell": f"C{row}",
                    "exclusion_reason": "source_credit_or_url_row",
                    "raw_value": statement,
                }
            )
            continue

        qsen_sequence += 1
        qsen_statement_id = f"QSEN_STMT_{qsen_sequence:04d}"
        domain_name = current_domain_raw.splitlines()[0].strip()
        domain_id = slug(domain_name, "qsen_domain")
        ksa_id = slug(current_ksa_raw, "ksa")
        domains[domain_id] = {
            "qsen_domain_id": domain_id,
            "qsen_domain_name": domain_name,
            "qsen_domain_raw": current_domain_raw,
        }
        ksas[ksa_id] = {
            "ksa_id": ksa_id,
            "ksa_name": current_ksa_raw,
            "ksa_raw": current_ksa_raw,
        }

        fact = {
            "qsen_statement_id": qsen_statement_id,
            "source_workbook": SOURCE_PATH.name,
            "source_sheet": "Entry Level",
            "source_row": row,
            "qsen_domain_id": domain_id,
            "qsen_domain_raw": current_domain_raw,
            "qsen_domain_name": domain_name,
            "ksa_id": ksa_id,
            "ksa_raw": current_ksa_raw,
            "qsen_statement_raw": statement,
            "mapping_cell_count_pre_split": 0,
            "bridge_row_count": 0,
            "mapping_status": "mapped",
            "needs_review": False,
            "review_reason": "",
        }
        facts.append(fact)

        for col, col_meta in entry_comp_columns.items():
            raw_mapping = cell(entry, row, col)
            if not raw_mapping:
                continue
            fact["mapping_cell_count_pre_split"] += 1
            for code in split_codes(raw_mapping):
                reasons = []
                if not SUBCOMPETENCY_RE.match(code):
                    reasons.append("malformed_subcompetency_code")
                parent_prefix = code[:-1] if SUBCOMPETENCY_RE.match(code) else ""
                competency_id = col_meta["competency_id"] or parent_prefix
                competency_name = col_meta["competency_name"]
                competency_raw = col_meta["competency_raw"]
                source_header_missing = bool(col_meta.get("missing_competency_header"))
                source_header_missing_resolved = False
                competency_title_source = "source_workbook_header"
                competency_title_source_url = ""
                supplemental = SUPPLEMENTAL_COMPETENCY_TITLES.get(competency_id)
                if col_meta.get("missing_competency_header"):
                    if supplemental:
                        competency_name = supplemental["competency_name"]
                        competency_title_source = supplemental["competency_title_source"]
                        competency_title_source_url = supplemental["competency_title_source_url"]
                        source_header_missing_resolved = True
                    else:
                        reasons.append("missing_competency_header")
                if parent_prefix and col_meta["competency_id"] and parent_prefix != col_meta["competency_id"]:
                    reasons.append("parent_prefix_mismatch")
                if competency_id and competency_id not in aacn_competencies:
                    aacn_competencies[competency_id] = {
                        "competency_id": competency_id,
                        "competency_name": competency_name,
                        "competency_raw": competency_raw,
                        "aacn_domain_id": col_meta["aacn_domain_id"],
                        "aacn_domain_number": col_meta["aacn_domain_number"],
                        "source_sheet": col_meta["source_sheet"],
                        "source_header_cell": col_meta["source_header_cell"],
                        "competency_title_source": competency_title_source,
                        "competency_title_source_url": competency_title_source_url,
                        "source_header_missing": source_header_missing,
                        "source_header_missing_resolved": source_header_missing_resolved,
                        "needs_review": bool(reasons),
                        "review_reason": "; ".join(reasons),
                    }
                bridges.append(
                    {
                        "bridge_id": f"BRIDGE_{len(bridges) + 1:05d}",
                        "qsen_statement_id": qsen_statement_id,
                        "source_workbook": SOURCE_PATH.name,
                        "source_sheet": "Entry Level",
                        "source_row": row,
                        "source_cell": f"{get_column_letter(col)}{row}",
                        "source_cell_raw": raw_mapping,
                        "aacn_domain_id": col_meta["aacn_domain_id"],
                        "aacn_domain_number": col_meta["aacn_domain_number"],
                        "aacn_domain_name": col_meta["aacn_domain_name"],
                        "aacn_domain_raw": col_meta["aacn_domain_raw"],
                        "competency_id": competency_id,
                        "competency_name": competency_name,
                        "competency_raw": competency_raw,
                        "competency_title_source": competency_title_source,
                        "competency_title_source_url": competency_title_source_url,
                        "source_header_missing": source_header_missing,
                        "source_header_missing_resolved": source_header_missing_resolved,
                        "subcompetency_code": code,
                        "needs_review": bool(reasons),
                        "review_reason": "; ".join(reasons),
                    }
                )

    duplicate_counter = Counter((b["qsen_statement_id"], b["subcompetency_code"]) for b in bridges)
    for bridge in bridges:
        if duplicate_counter[(bridge["qsen_statement_id"], bridge["subcompetency_code"])] > 1:
            bridge["needs_review"] = True
            bridge["review_reason"] = "; ".join(
                r for r in [bridge["review_reason"], "duplicate_statement_subcompetency_mapping"] if r
            )

    bridge_by_statement = defaultdict(list)
    for bridge in bridges:
        bridge_by_statement[bridge["qsen_statement_id"]].append(bridge)

    fact_by_id = {fact["qsen_statement_id"]: fact for fact in facts}
    for fact in facts:
        related = bridge_by_statement.get(fact["qsen_statement_id"], [])
        fact["bridge_row_count"] = len(related)
        if not related:
            if is_declared_no_alignment(fact["qsen_statement_raw"]):
                fact["mapping_status"] = "intentionally_unmapped_source_declared"
            else:
                fact["mapping_status"] = "unmapped_needs_review"
                fact["needs_review"] = True
                fact["review_reason"] = "no_aacn_mapping"
                review_queue.append(
                    {
                        "review_id": f"REV_{len(review_queue) + 1:05d}",
                        "record_type": "Fact_QSEN_Statements",
                        "qsen_statement_id": fact["qsen_statement_id"],
                        "source_sheet": fact["source_sheet"],
                        "source_row": fact["source_row"],
                        "source_cell": f"C{fact['source_row']}",
                        "review_reason": "no_aacn_mapping",
                        "raw_value": fact["qsen_statement_raw"],
                    }
                )
        elif any(b["needs_review"] for b in related):
            fact["mapping_status"] = "mapped_needs_review"
            fact["needs_review"] = True
            fact["review_reason"] = "; ".join(sorted({b["review_reason"] for b in related if b["review_reason"]}))

    for bridge in bridges:
        if bridge["needs_review"]:
            review_queue.append(
                {
                    "review_id": f"REV_{len(review_queue) + 1:05d}",
                    "record_type": "Bridge_QSEN_AACN",
                    "qsen_statement_id": bridge["qsen_statement_id"],
                    "bridge_id": bridge["bridge_id"],
                    "source_sheet": bridge["source_sheet"],
                    "source_row": bridge["source_row"],
                    "source_cell": bridge["source_cell"],
                    "review_reason": bridge["review_reason"],
                    "raw_value": bridge["source_cell_raw"],
                }
            )

    master = []
    for bridge in bridges:
        fact = fact_by_id[bridge["qsen_statement_id"]]
        master.append(
            {
                "qsen_statement_id": fact["qsen_statement_id"],
                "qsen_domain_name": fact["qsen_domain_name"],
                "ksa": fact["ksa_raw"],
                "qsen_statement_raw": fact["qsen_statement_raw"],
                "aacn_domain_number": bridge["aacn_domain_number"],
                "aacn_domain_name": bridge["aacn_domain_name"],
                "competency_id": bridge["competency_id"],
                "competency_name": bridge["competency_name"],
                "subcompetency_code": bridge["subcompetency_code"],
                "competency_title_source": bridge["competency_title_source"],
                "competency_title_source_url": bridge["competency_title_source_url"],
                "source_header_missing": bridge["source_header_missing"],
                "source_header_missing_resolved": bridge["source_header_missing_resolved"],
                "source_sheet": bridge["source_sheet"],
                "source_row": bridge["source_row"],
                "source_cell": bridge["source_cell"],
                "source_cell_raw": bridge["source_cell_raw"],
                "needs_review": bridge["needs_review"],
                "review_reason": bridge["review_reason"],
            }
        )

    coverage_rows = []
    by_domain = defaultdict(lambda: {"statement_ids": set(), "bridge_count": 0, "review_count": 0})
    by_ksa = defaultdict(lambda: {"statement_ids": set(), "bridge_count": 0, "review_count": 0})
    by_aacn_domain = defaultdict(lambda: {"statement_ids": set(), "bridge_count": 0, "review_count": 0})
    by_competency = defaultdict(lambda: {"statement_ids": set(), "bridge_count": 0, "review_count": 0})
    by_review_reason = Counter()

    for fact in facts:
        by_domain[fact["qsen_domain_name"]]["statement_ids"].add(fact["qsen_statement_id"])
        by_ksa[fact["ksa_raw"]]["statement_ids"].add(fact["qsen_statement_id"])
        if fact["needs_review"] and fact["review_reason"]:
            for reason in fact["review_reason"].split("; "):
                by_review_reason[reason] += 1
    for bridge in bridges:
        fact = fact_by_id[bridge["qsen_statement_id"]]
        by_domain[fact["qsen_domain_name"]]["bridge_count"] += 1
        by_ksa[fact["ksa_raw"]]["bridge_count"] += 1
        by_aacn_domain[f"{bridge['aacn_domain_number']} {bridge['aacn_domain_name']}"]["statement_ids"].add(
            bridge["qsen_statement_id"]
        )
        by_aacn_domain[f"{bridge['aacn_domain_number']} {bridge['aacn_domain_name']}"]["bridge_count"] += 1
        by_competency[f"{bridge['competency_id']} {bridge['competency_name']}"]["statement_ids"].add(
            bridge["qsen_statement_id"]
        )
        by_competency[f"{bridge['competency_id']} {bridge['competency_name']}"]["bridge_count"] += 1
        if bridge["needs_review"]:
            by_domain[fact["qsen_domain_name"]]["review_count"] += 1
            by_ksa[fact["ksa_raw"]]["review_count"] += 1
            by_aacn_domain[f"{bridge['aacn_domain_number']} {bridge['aacn_domain_name']}"]["review_count"] += 1
            by_competency[f"{bridge['competency_id']} {bridge['competency_name']}"]["review_count"] += 1

    def add_coverage(kind: str, data: dict):
        for label, row in sorted(data.items()):
            coverage_rows.append(
                {
                    "coverage_type": kind,
                    "label": label,
                    "statement_count": len(row["statement_ids"]),
                    "bridge_row_count": row["bridge_count"],
                    "review_count": row["review_count"],
                }
            )

    add_coverage("qsen_domain", by_domain)
    add_coverage("ksa", by_ksa)
    add_coverage("aacn_domain", by_aacn_domain)
    add_coverage("aacn_competency", by_competency)
    for reason, count in sorted(by_review_reason.items()):
        coverage_rows.append(
            {
                "coverage_type": "review_reason",
                "label": reason,
                "statement_count": "",
                "bridge_row_count": "",
                "review_count": count,
            }
        )

    domain_sheet_mapped = sum(
        sheet["mapped_cells_pre_split"] for sheet in inventory["sheets"] if sheet["sheet_name"] != "Entry Level"
    )
    entry_sheet = next(sheet for sheet in inventory["sheets"] if sheet["sheet_name"] == "Entry Level")
    reconciliation = {
        "sheet_count_ok": inventory["sheet_count"] == 11,
        "entry_level_data_rows_ok": entry_sheet["data_rows_from_6"] == 214,
        "entry_level_mapped_cells_pre_split_ok": entry_sheet["mapped_cells_pre_split"] == 651,
        "domain_sheet_mapped_cells_sum": domain_sheet_mapped,
        "domain_reconciliation_ok": domain_sheet_mapped == entry_sheet["mapped_cells_pre_split"] == 651,
    }

    return {
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_path": str(SOURCE_PATH),
            "source_hash": inventory["source_hash"],
            "source_workbook": SOURCE_PATH.name,
            "supplemental_sources": [
                {
                    "use": "Resolved missing source workbook parent competency title for 1.1 mappings in Entry Level column D.",
                    **source,
                }
                for source in SUPPLEMENTAL_COMPETENCY_TITLES.values()
            ],
        },
        "inventory": inventory,
        "reconciliation": reconciliation,
        "dim_qsen_domain": sorted(domains.values(), key=lambda x: x["qsen_domain_name"]),
        "dim_ksa": sorted(ksas.values(), key=lambda x: x["ksa_name"]),
        "dim_aacn_domain": sorted(aacn_domains.values(), key=lambda x: int(x["aacn_domain_number"] or 999)),
        "dim_aacn_competency": sorted(
            aacn_competencies.values(),
            key=lambda x: [int(part) for part in x["competency_id"].split(".")],
        ),
        "fact_qsen_statements": facts,
        "bridge_qsen_aacn": bridges,
        "master_canonical": master,
        "coverage_summary": coverage_rows,
        "review_queue": review_queue,
        "source_exclusions": source_exclusions,
    }


def main() -> int:
    INTERMEDIATE_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE_PATH, read_only=False, data_only=False)
    inventory = workbook_inventory(wb)
    data = build_crosswalk_data(wb, inventory)
    SOURCE_INVENTORY_PATH.write_text(json.dumps(inventory, indent=2), encoding="utf-8")
    CROSSWALK_DATA_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(json.dumps({
        "source_inventory": str(SOURCE_INVENTORY_PATH),
        "crosswalk_data": str(CROSSWALK_DATA_PATH),
        "sheet_count": inventory["sheet_count"],
        "entry_level_rows": next(s for s in inventory["sheets"] if s["sheet_name"] == "Entry Level")["data_rows_from_6"],
        "entry_level_mapped_cells_pre_split": next(s for s in inventory["sheets"] if s["sheet_name"] == "Entry Level")["mapped_cells_pre_split"],
        "bridge_rows": len(data["bridge_qsen_aacn"]),
        "review_queue_rows": len(data["review_queue"]),
        "source_exclusion_rows": len(data["source_exclusions"]),
        "reconciliation": data["reconciliation"],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
